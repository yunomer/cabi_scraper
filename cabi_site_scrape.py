import urllib.request
import os
import io
import re
import sys
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# This displays the progress bar of downloading
def updt(total, progress):
    """
    Displays or updates a console progress bar.

    Original source: https://stackoverflow.com/a/15860757/1391441
    """
    barLength, status = 20, ""
    progress = float(progress) / float(total)
    if progress >= 1.:
        progress, status = 1, "\r\n"
    block = int(round(barLength * progress))
    text = "\r[{}] {:.0f}% {}".format(
        "#" * block + "-" * (barLength - block), round(progress * 100, 0),
        status)
    sys.stdout.write(text)
    sys.stdout.flush()


# This function creates the file where the downloaded HTML will be places
def create_file(filename):
    cwd = os.getcwd()
    final_directory = os.path.join(cwd, filename)
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)


# This function opens the link, saves the HTML in the proper directory and then returns a BS4 object
def open_save_link(link, filename):
    filename = filename.rstrip()
    filename = re.sub('[^A-Za-z0-9 ]+', ' ', filename)
    filename = filename.rstrip()
    page = urllib.request.urlopen(link)
    page = BeautifulSoup(page, "lxml")
    page = page.prettify()
    create_file(filename)
    directory = os.getcwd()
    save_file = os.path.join(directory, filename)
    save_file = os.path.join(save_file, filename + ".html")
    with io.open(save_file, "w+", encoding="utf-8") as f:
        f.write(page)
        f.close()
    return page


# This function strips data from the site and places it in the new Sheet.
def strip_n_save(bs4_object, index, database, sheet):

    species_name = database.cell(row=index, column=1).value
    soup = BeautifulSoup(bs4_object, "html.parser")
    index = sheet.max_row+1

    try:
        identity = soup.find("div", {"id": "toidentity"})
        identity = str(identity).strip()
        identity = re.sub("\n+", " ", identity.strip())
        identity = " ".join(identity.split())
    except:
        pass
    try:
        host_plants = soup.find("div", {"id": "tohostPlants"})
        host_plants = str(host_plants).strip()
        host_plants = re.sub("\n+", " ", host_plants.strip())
        host_plants = " ".join(host_plants.split())
    except:
        pass
    try:
        natural_enemies = soup.find("div", {"id": "tonaturalEnemies"})
        natural_enemies = str(natural_enemies).strip()
        natural_enemies = re.sub("\n+", " ", natural_enemies.strip())
        natural_enemies = " ".join(natural_enemies.split())
    except:
        pass
    try:
        pathway_vectors = soup.find("div", {"id": "topathwayVectors"})
        pathway_vectors = str(pathway_vectors).strip()
        pathway_vectors = re.sub("\n+", " ", pathway_vectors.strip())
        pathway_vectors = " ".join(pathway_vectors.split())
    except:
        pass
    try:
        signs_symptoms = soup.find("div", {"id": "tosymptomsOrSigns"})
        signs_symptoms = str(signs_symptoms).strip()
        signs_symptoms = re.sub("\n+", " ", signs_symptoms.strip())
        signs_symptoms = " ".join(signs_symptoms.split())
    except:
        pass

    sheet.cell(row=index, column=1).value = species_name
    sheet.cell(row=index, column=2).value = identity
    sheet.cell(row=index, column=3).value = str(host_plants)
    sheet.cell(row=index, column=4).value = str(natural_enemies)
    sheet.cell(row=index, column=5).value = str(pathway_vectors)
    sheet.cell(row=index, column=6).value = str(signs_symptoms)


# load the workbook
wb = load_workbook(filename="cabi_database.xlsx", data_only=True)

# Select sheets to work with
database = wb['links']
new_sheet = wb.create_sheet(title="data")

# number of lines
total_rows = database.max_row

new_sheet.cell(row=1, column=1).value = "Species Name"
new_sheet.cell(row=1, column=2).value = "Identity Data"
new_sheet.cell(row=1, column=3).value = "Host Plant Data"
new_sheet.cell(row=1, column=4).value = "Natural Enemies"
new_sheet.cell(row=1, column=5).value = "Pathway Vectors"
new_sheet.cell(row=1, column=6).value = "Signs or Symptoms"

for row in range(1, total_rows):
    link = database.cell(row=row, column=2).value.split(';')
    site_data = open_save_link(link[0], database.cell(row=row, column=1).value)
    strip_n_save(site_data, row, database, new_sheet)
    # this is progress bar
    time.sleep(.1)
    updt(total_rows-1, row)

wb.save("cabi_database.xlsx")
