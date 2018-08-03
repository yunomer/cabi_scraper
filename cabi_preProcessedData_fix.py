from openpyxl import load_workbook
from bs4 import BeautifulSoup
import sys
import time


# This displays the progress bar of downloading
def updt(total, progress):
    """
    Displays or updates a console progress bar.

    Original source: https://stackoverflow.com/a/15860757/1391441
    """
    barLength, status = 20, ""
    progress = float(progress) / float(total)
    if progress >= 1.:
        progress, status = 1, "\r\n"
    block = int(round(barLength * progress))
    text = "\r[{}] {:.0f}% {}".format(
        "#" * block + "-" * (barLength - block), round(progress * 100, 0),
        status)
    sys.stdout.write(text)
    sys.stdout.flush()


def process_sp(cell):
    return cell


def process_id(cell):
    dictionary = []
    if cell == "None":
        return []
    soup = BeautifulSoup(cell, "html.parser")
    content = soup.find_all("h4")
    for i in range(0, len(content)):
        child_dictionary = []
        key = content[i].string
        value = []
        list_elements = content[i].findNext('ul').findAll('li')
        for j in range(0, len(list_elements)):
            try:
                value.append(list_elements[j].string)
            except Exception as e:
                print(e)
                value.append([""])
        try:
            child_dictionary.append(key)
            child_dictionary.append(value)
            dictionary.append(child_dictionary)
        except Exception as e:
            print(e)
            dictionary.append([])
    return dictionary


def process_ph(cell):
    dictionary = []
    if cell == "None":
        return []
    soup = BeautifulSoup(cell, "html.parser")
    content = soup.find_all("tr")
    try:
        for pest in range(0, len(content)):
            child_dictionary = []
            name = []
            family = []
            child_content = content[pest].findAll('td')
            try:
                temp = child_content[0].find("a")
                temp = temp.string
            except:
                temp = child_content[0].string
                pass
            name.append(temp)
            family.append(child_content[1].string)
            child_dictionary.append(name)
            child_dictionary.append(family)
            dictionary.append(child_dictionary)
    except Exception as e:
        print(e)
        print(soup)
        exit(0)
    return dictionary


def place_text(sheet, name, identity_data, host_data):
    max_row = sheet.max_row + 1

    try:
        data_piece = identity_data[0][0]
        if data_piece.strip() == "Preferred Scientific Name":
            psn = ', '.join(identity_data[0][1])
        else:
            psn = "N/A"
    except:
        psn = "N/A"
        pass

    try:
        data_piece = identity_data[1][0]
        if data_piece.strip() == "Preferred Common Name":
            pcn = ', '.join(identity_data[1][1])
        else:
            pcn = "N/A"
    except:
        pcn = "N/A"
        pass

    try:
        data_piece = identity_data[2][0]
        if data_piece.strip() == "Other Scientific Names":
            osn = ', '.join(identity_data[2][1])
        else:
            osn = "N/A"
    except:
        osn = "N/A"
        pass

    if len(host_data) > 0:
        for row in range(0, len(host_data)):
            child_name = host_data[row][0][0]
            family = host_data[row][1][0]

            sheet.cell(row=max_row, column=1).value = name
            sheet.cell(row=max_row, column=2).value = psn
            sheet.cell(row=max_row, column=3).value = pcn
            sheet.cell(row=max_row, column=4).value = osn
            sheet.cell(row=max_row, column=5).value = family
            sheet.cell(row=max_row, column=6).value = child_name
            max_row += 1
    else:
        sheet.cell(row=max_row, column=1).value = name
        sheet.cell(row=max_row, column=2).value = psn
        sheet.cell(row=max_row, column=3).value = pcn
        sheet.cell(row=max_row, column=4).value = osn
        sheet.cell(row=max_row, column=5).value = "N/A"
        sheet.cell(row=max_row, column=6).value = "N/A"
        max_row += 1


def main():
    # load the workbook
    wb = load_workbook(filename="cabi_database.xlsx", data_only=True)

    # Select sheets to work with
    database = wb['data']
    new_sheet = wb.create_sheet(title="processed")

    # number of lines
    total_rows = database.max_row

    new_sheet.cell(row=1, column=1).value = "Species Name"
    new_sheet.cell(row=1, column=2).value = "Scientific Name"
    new_sheet.cell(row=1, column=3).value = "Common name"
    new_sheet.cell(row=1, column=4).value = "Other Scientific Names"
    new_sheet.cell(row=1, column=5).value = "Family Name"
    new_sheet.cell(row=1, column=6).value = "Host Name"

    for row in range(2, total_rows+1):
        species_name = process_sp(database.cell(row=row, column=1).value)
        identity_data = process_id(database.cell(row=row, column=2).value)
        plant_host = process_ph(database.cell(row=row, column=3).value)
        place_text(new_sheet, species_name, identity_data, plant_host)
        time.sleep(.1)
        updt(total_rows - 1, row)

    wb.save("cabi_database.xlsx")


if __name__ == '__main__':
    main()
