import urllib.request
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# load the workbook
wb = load_workbook(filename="cabi_database.xlsx", data_only=True)

# Select the Sheet to work with
rawData = wb['sheet1']
linkSheet = wb.create_sheet(title="links")

# data = [pair, pair,...]
data = []
pair = []

row_number = 1
total_rows = 5129
next_column = 3

while row_number < total_rows:
    try:
        name = rawData.cell(row=row_number, column=1).value.split("\t")
        rawString = rawData.cell(row=row_number, column=2).value.split("\t")
        links = rawString[len(rawString)-1]
        link_check = links.split("://")
        while link_check[0] != 'https':
            rawString = rawData.cell(row=row_number, column=next_column).value.split("\t")
            links = rawString[len(rawString)-1]
            link_check = links.split("://")
            if next_column is not 7:
                next_column += 1
            else:
                next_column = 3
                break
        next_column = 3
        pair.append(name[0])
        pair.append(links)
        data.append(pair)
        pair = []
        row_number += 1
    except Exception as e:
        print(e)
        exit(0)
        pass

row_number = 1
for i in range(0, len(data)):
    link_to_save = data[i]
    linkSheet.cell(row=row_number, column=1).value = link_to_save[0]
    linkSheet.cell(row=row_number, column=2).value = link_to_save[1]
    row_number += 1

wb.save("cabi_database.xlsx")
